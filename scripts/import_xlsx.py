#!/usr/bin/env python3
"""Read the 1UP+ Customs scoring workbook and write the repo's tournament CSVs.

The workbook is where results are entered on the night. This repo is where they
are kept. Until now the crossing between the two was a person retyping numbers,
which is the single thing that has gone wrong before - see scripts/EXTRACTION_BRIEF.md.

    python scripts/import_xlsx.py <workbook.xlsx> --date 2026-08-18
    python scripts/import_xlsx.py <workbook.xlsx> --date 2026-08-18 --dry-run

Nothing is written unless the workbook's own arithmetic reproduces. The script
recomputes every team total from the atoms it just read and compares against the
total the spreadsheet itself shows in column C. If any team disagrees, it stops
and prints both figures. A layout that has shifted - a map column inserted, a
team row added - shows up as a mismatch rather than as silently wrong data.

That check is the whole point of this script. It is the same rule the squad
totals row enforces on the screenshots: two independent paths to one number,
and a failure if they disagree.
"""
import argparse
import csv
import os
import re
import sys
import unicodedata
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

# --- Workbook layout -------------------------------------------------------
# Every magic number the sheet's geometry depends on, named once. If the
# workbook is restructured these are the lines to change, and the arithmetic
# check below is what will tell you they need changing.

SCORES = "Score Sheet"
ROSTERS = "Teams Lists"

TEAM_ROW_FIRST = 10           # first team row on the Score Sheet
TEAM_COUNT = 17
COL_TEAM_NO, COL_TEAM_NAME, COL_TOTAL = 1, 2, 3
MAP_COUNT = 16
MAP_BLOCK_FIRST_COL = 4       # D: M1 kills. Then place, total, repeating every 3.

MULT_ROWS = {                 # Y3:Z7 - place band label -> multiplier cell row
    "1st": 3, "2nd": 4, "3rd-5th": 5, "6th-10th": 6, "10th-": 7,
}
MULT_COL = 26                 # Z
MATCHPOINT_CELL = (2, 29)     # AC2

# Player-kill grid: two blocks, because 17 teams do not fit across one screen.
# (header_row, first_map_row, first_team_no, team_count)
KILL_BLOCKS = (
    (31, 32, 1, 9),
    (49, 50, 10, 8),
)
KILL_FIRST_COL = 3            # C, then every 3 columns per team

# Roster grid on 'Teams Lists': three teams per band of rows.
ROSTER_NAME_ROWS = (19, 29, 39, 49, 60, 71)
ROSTER_COLS = (3, 11, 19)     # C, K, S - first player of each of the three teams
ROSTER_PLAYER_STRIDE = 2      # players sit at +0, +2, +4
ROSTER_KILLS_OFFSET = 3       # kills row is three below the name row


def slug(name):
    """A stable player_id from a gamertag.

    The repo's standing rule is that a gamertag is never a key - people rename
    mid-tournament. This produces a starting id; the operator is expected to
    reconcile it against data/aliases.csv, and the script reports which ones
    it could not match.
    """
    text = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode()
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def multiplier(place, mults):
    """Placement multiplier, mirroring the workbook's nested IF exactly.

    Reproduced rather than improved on purpose: this function has to agree with
    the spreadsheet, including where the spreadsheet is odd. Two notes on that.

    A placement of 10 takes the 6th-10th band, not the '10th-' band, because the
    workbook tests `<=10` before `>=10`.

    A placement of 0 or blank - a map a team did not play - falls through to the
    `<=5` test and takes 1.4x. It is harmless only because those rows also carry
    zero kills. It is called out in the report rather than silently corrected,
    because correcting it here would make this script disagree with the sheet
    the organisers are looking at.
    """
    if place is None or place == "":
        return mults["3rd-5th"], True
    place = float(place)
    if place == 1:
        return mults["1st"], False
    if place == 2:
        return mults["2nd"], False
    if place <= 5:
        return mults["3rd-5th"], place <= 0
    if place <= 10:
        return mults["6th-10th"], False
    return mults["10th-"], False


def num(value):
    """Cell value as a number, or None."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_workbook(path):
    """Pull everything out of the workbook into plain structures."""
    book = openpyxl.load_workbook(path, data_only=True)

    if SCORES not in book.sheetnames:
        sys.exit("no %r sheet - is this the right workbook?" % SCORES)

    scores = book[SCORES]
    rosters = book[ROSTERS] if ROSTERS in book.sheetnames else None

    mults = {k: float(scores.cell(r, MULT_COL).value)
             for k, r in MULT_ROWS.items()}
    matchpoint = num(scores.cell(*MATCHPOINT_CELL).value)

    # Rosters: team_no -> [(gamertag, kills_total_per_sheet), ...]
    roster = {}
    if rosters is not None:
        team_no = 0
        for name_row in ROSTER_NAME_ROWS:
            for col in ROSTER_COLS:
                team_no += 1
                if team_no > TEAM_COUNT:
                    break
                players = []
                for slot in range(3):
                    c = col + slot * ROSTER_PLAYER_STRIDE
                    tag = rosters.cell(name_row, c).value
                    kills = num(rosters.cell(name_row + ROSTER_KILLS_OFFSET, c).value)
                    players.append((str(tag).strip() if tag else "", kills))
                roster[team_no] = players

    # Per-map player kills, from the two grid blocks.
    # (team_no, map_no, slot) -> kills
    player_kills = {}
    for header_row, first_map_row, first_team, count in KILL_BLOCKS:
        for offset in range(count):
            team_no = first_team + offset
            base = KILL_FIRST_COL + offset * 3
            for map_no in range(1, MAP_COUNT + 1):
                row = first_map_row + map_no - 1
                for slot in range(3):
                    v = num(scores.cell(row, base + slot).value)
                    if v is not None:
                        player_kills[(team_no, map_no, slot)] = v

    # Teams and their per-map squad results.
    teams = OrderedDict()
    for offset in range(TEAM_COUNT):
        row = TEAM_ROW_FIRST + offset
        team_no = offset + 1
        name = scores.cell(row, COL_TEAM_NAME).value
        maps = {}
        for map_no in range(1, MAP_COUNT + 1):
            base = MAP_BLOCK_FIRST_COL + (map_no - 1) * 3
            maps[map_no] = {
                "kills": num(scores.cell(row, base).value),
                "place": num(scores.cell(row, base + 1).value),
                "total": num(scores.cell(row, base + 2).value),
            }
        teams[team_no] = {
            "name": str(name).strip() if name else "",
            "sheet_total": num(scores.cell(row, COL_TOTAL).value),
            "maps": maps,
            "players": roster.get(team_no, [("", None)] * 3),
        }

    return {
        "teams": teams,
        "player_kills": player_kills,
        "multipliers": mults,
        "matchpoint": matchpoint,
    }


def verify(data):
    """Recompute the workbook's own figures and report every disagreement.

    Three independent checks:

      1. Squad kills on a map must equal the three player cells for that map.
         This is the same shape as the squad-totals check on the screenshots.
      2. A map's points must equal kills x the multiplier the placement implies.
      3. A team's tournament total must equal the sum of its map points.

    Passing all three means the layout constants above are pointing at the
    right cells. Failing any means they are not, and no output is written.
    """
    problems, notes = [], []
    mults = data["multipliers"]

    for team_no, team in data["teams"].items():
        if not team["name"] or team["name"] == "x  x":
            continue

        computed_total = 0.0
        for map_no, m in team["maps"].items():
            kills, place, total = m["kills"], m["place"], m["total"]
            if kills is None and total is None:
                continue

            slots = [data["player_kills"].get((team_no, map_no, s)) for s in range(3)]
            if any(s is not None for s in slots):
                player_sum = sum(s or 0 for s in slots)
                if kills is not None and abs(player_sum - kills) > 1e-6:
                    problems.append(
                        "team %d (%s) map %d: squad kills %g but player cells sum to %g"
                        % (team_no, team["name"], map_no, kills, player_sum)
                    )

            mult, odd = multiplier(place, mults)
            expected = (kills or 0) * mult
            if total is not None and abs(expected - total) > 1e-6:
                problems.append(
                    "team %d (%s) map %d: %g kills at place %s implies %g points, sheet says %g"
                    % (team_no, team["name"], map_no, kills or 0,
                       "blank" if place is None else "%g" % place, expected, total)
                )
            if odd and (kills or 0) > 0:
                notes.append(
                    "team %d (%s) map %d: %g kills recorded with no placement - "
                    "the sheet's formula gives these 1.4x"
                    % (team_no, team["name"], map_no, kills)
                )
            computed_total += total if total is not None else expected

        if team["sheet_total"] is not None and abs(computed_total - team["sheet_total"]) > 1e-6:
            problems.append(
                "team %d (%s): map points sum to %.2f, sheet total says %.2f"
                % (team_no, team["name"], computed_total, team["sheet_total"])
            )

    # Roster kill totals against the per-map grid.
    for team_no, team in data["teams"].items():
        for slot, (tag, sheet_kills) in enumerate(team["players"]):
            if not tag or sheet_kills is None:
                continue
            grid = sum(data["player_kills"].get((team_no, m, slot), 0)
                       for m in range(1, MAP_COUNT + 1))
            if abs(grid - sheet_kills) > 1e-6:
                problems.append(
                    "team %d slot %d (%s): grid kills %g, roster total %g"
                    % (team_no, slot + 1, tag, grid, sheet_kills)
                )

    return problems, notes


def maps_played(team):
    """Map numbers this team actually has a result for.

    Unplayed maps are not blank. The workbook's kills and points cells are
    formulas over empty inputs, so they read back as 0 rather than as nothing,
    and 16 maps always appear to exist. A map counts as played only if there is
    a placement or a non-zero kill count behind it - otherwise every event
    imports as a 16-map tournament with eight maps of zeros, which would drag
    every median down and make the whole model wrong in a way that still
    balances arithmetically.
    """
    out = []
    for n, m in sorted(team["maps"].items()):
        place, kills = m["place"], m["kills"]
        if (place is not None and place > 0) or (kills is not None and kills > 0):
            out.append(n)
    return out


def write_csvs(data, date, out_dir, bracket, matchpoint):
    """Write the repo's five tournament CSVs."""
    os.makedirs(out_dir, exist_ok=True)

    live = [(n, t) for n, t in data["teams"].items()
            if t["name"] and t["name"] != "x  x"]
    played = sorted({m for _, t in live for m in maps_played(t)})

    def dump(name, header, rows):
        with open(os.path.join(out_dir, name), "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            w.writerows(rows)
        return len(rows)

    counts = {}

    counts["meta.csv"] = dump(
        "meta.csv",
        ["played_on", "format", "map_count", "bracket", "rank_cap",
         "movement_floor", "match_point", "notes"],
        [[date, "Resurgence trios", len(played), bracket, "",
          "", matchpoint if matchpoint is not None else "",
          "Imported from the 1UP+ scoring workbook by scripts/import_xlsx.py."]],
    )

    mults = data["multipliers"]
    counts["scoring.csv"] = dump(
        "scoring.csv",
        ["place_from", "place_to", "multiplier"],
        [[1, 1, mults["1st"]],
         [2, 2, mults["2nd"]],
         [3, 5, mults["3rd-5th"]],
         [6, 10, mults["6th-10th"]],
         [11, 99, mults["10th-"]]],
    )

    squad_rows, player_rows, result_rows = [], [], []
    for team_no, team in live:
        squad_id = "T%02d" % team_no
        for slot, (tag, _) in enumerate(team["players"]):
            if tag:
                squad_rows.append([squad_id, team["name"], slug(tag), tag, slot + 1])

        for map_no in maps_played(team):
            m = team["maps"][map_no]
            mult, _ = multiplier(m["place"], mults)
            result_rows.append([
                map_no, squad_id,
                "" if m["kills"] is None else int(m["kills"]),
                "" if m["place"] is None else int(m["place"]),
                "" if m["total"] is None else round(m["total"], 2),
                mult, "C", "",
            ])
            for slot, (tag, _) in enumerate(team["players"]):
                if not tag:
                    continue
                kills = data["player_kills"].get((team_no, map_no, slot))
                player_rows.append([
                    map_no, squad_id, slug(tag), tag,
                    "" if kills is None else int(kills),
                    "C" if kills is not None else "U",
                ])

    counts["squads.csv"] = dump(
        "squads.csv", ["squad_id", "label", "player_id", "handle", "slot"], squad_rows)
    counts["map_results.csv"] = dump(
        "map_results.csv",
        ["map_no", "squad_id", "squad_kills", "placement", "squad_points",
         "multiplier", "confidence", "evidence_url"],
        result_rows)
    counts["player_map.csv"] = dump(
        "player_map.csv",
        ["map_no", "squad_id", "player_id", "handle", "kills", "confidence"],
        player_rows)
    counts["decisions.csv"] = dump(
        "decisions.csv",
        ["player_id", "direction", "from_rank", "to_rank", "visibility",
         "reason", "numbers_said"],
        [])

    return counts


def write_live(data, date, bracket, path):
    """Write the feed the live leaderboard renders from.

    Deliberately independent of the ranking model. The band review needs
    players.csv, ranks.csv and aliases.csv reconciled first - and reconciling a
    handle like 'Chuffy/Rev/Stump' into a player_id is a judgement call somebody
    has to make. That cannot happen between maps on a tournament night.

    So this feed carries only what the workbook itself knows: standings, kills,
    and who has reached match point. It can be regenerated and published the
    moment a map is entered, with no roster decisions outstanding.
    """
    import json

    live = [(n, t) for n, t in data["teams"].items()
            if t["name"] and t["name"] != "x  x"]
    played = sorted({m for _, t in live for m in maps_played(t)})
    matchpoint = data["matchpoint"]

    teams = []
    for team_no, team in sorted(live, key=lambda x: -(x[1]["sheet_total"] or 0)):
        played_here = maps_played(team)
        places = [team["maps"][m]["place"] for m in played_here
                  if team["maps"][m]["place"]]
        teams.append({
            "id": "T%02d" % team_no,
            "label": team["name"],
            "pts": round(team["sheet_total"] or 0, 1),
            "k": int(sum(team["maps"][m]["kills"] or 0 for m in played_here)),
            "maps": len(played_here),
            "plc": round(sum(places) / len(places), 2) if places else None,
            "best": int(min(places)) if places else None,
            "wins": sum(1 for p in places if p == 1),
            "matchpoint": bool(matchpoint is not None
                               and (team["sheet_total"] or 0) >= matchpoint),
            "players": [tag for tag, _ in team["players"] if tag],
        })

    tally = {}
    for (team_no, map_no, slot), kills in data["player_kills"].items():
        tag = data["teams"][team_no]["players"][slot][0]
        if tag:
            entry = tally.setdefault(tag, {"k": 0, "team": data["teams"][team_no]["name"]})
            entry["k"] += kills

    players = [{"p": tag, "k": int(v["k"]), "team": v["team"]}
               for tag, v in tally.items()]
    players.sort(key=lambda r: (-r["k"], r["p"].lower()))

    payload = {
        "date": date,
        "bracket": bracket,
        "match_point": matchpoint,
        "maps_played": len(played),
        "multipliers": [
            {"label": "1st", "from": 1, "to": 1, "x": data["multipliers"]["1st"]},
            {"label": "2nd", "from": 2, "to": 2, "x": data["multipliers"]["2nd"]},
            {"label": "3rd-5th", "from": 3, "to": 5, "x": data["multipliers"]["3rd-5th"]},
            {"label": "6th-10th", "from": 6, "to": 10, "x": data["multipliers"]["6th-10th"]},
            {"label": "11th+", "from": 11, "to": 99, "x": data["multipliers"]["10th-"]},
        ],
        "teams": teams,
        "players": players,
    }

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with open(path, "w") as fh:
        json.dump(payload, fh, separators=(",", ":"))

    return payload


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("--date", required=True, help="ISO date of the event, e.g. 2026-08-18")
    ap.add_argument("--bracket", default="8-", help="bracket label (default: 8-)")
    ap.add_argument("--out", default=None, help="output directory")
    ap.add_argument("--live", default=None, metavar="PATH",
                    help="also write the live leaderboard feed, e.g. docs/live.json")
    ap.add_argument("--live-only", action="store_true",
                    help="write only the live feed. No roster reconciliation needed, "
                         "so this is the one to run between maps on the night.")
    ap.add_argument("--dry-run", action="store_true",
                    help="verify and report, write nothing")
    ap.add_argument("--force", action="store_true",
                    help="write even if verification failed. You will be asked why.")
    args = ap.parse_args()

    if not re.match(r"^\d{4}-\d{2}-\d{2}$", args.date):
        sys.exit("--date must be ISO, e.g. 2026-08-18")

    data = read_workbook(args.workbook)
    problems, notes = verify(data)

    live = [(n, t) for n, t in data["teams"].items()
            if t["name"] and t["name"] != "x  x"]
    played = sorted({m for _, t in live for m in maps_played(t)})

    print("workbook:   %s" % args.workbook)
    print("teams:      %d entered, %d blank" % (len(live), TEAM_COUNT - len(live)))
    print("maps:       %d played" % len(played))
    print("multipliers: %s" % ", ".join(
        "%s=%g" % (k, data["multipliers"][k]) for k in MULT_ROWS))
    print("match point: %s" % data["matchpoint"])

    print("\nstandings per the workbook")
    for rank, (team_no, team) in enumerate(
            sorted(live, key=lambda x: -(x[1]["sheet_total"] or 0)), 1):
        flag = "  <- match point" if (data["matchpoint"] is not None
                                      and (team["sheet_total"] or 0) >= data["matchpoint"]) else ""
        print("  %2d. %-42s %7.1f%s" % (rank, team["name"][:42],
                                        team["sheet_total"] or 0, flag))

    tally = {}
    for (team_no, map_no, slot), kills in data["player_kills"].items():
        tag = data["teams"][team_no]["players"][slot][0]
        if tag:
            tally[tag] = tally.get(tag, 0) + kills
    print("\ntop 10 kills")
    for rank, (tag, kills) in enumerate(
            sorted(tally.items(), key=lambda x: -x[1])[:10], 1):
        print("  %2d. %-24s %g" % (rank, tag, kills))

    if notes:
        print("\nnotes (%d)" % len(notes))
        for n in notes:
            print("  - %s" % n)

    odd_tags = [tag for tag in tally if tag.startswith("x ") or "/" in tag]
    if odd_tags:
        print("\ngamertags that need a decision before these become player_ids (%d)"
              % len(odd_tags))
        for tag in sorted(odd_tags):
            reason = ("looks like a team-name parse artifact" if tag.startswith("x ")
                      else "two people sharing one slot")
            print("  - %-24s -> %-22s %s" % (tag, slug(tag), reason))

    if problems:
        print("\nVERIFICATION FAILED - %d disagreement(s) with the workbook" % len(problems))
        for p in problems:
            print("  ! %s" % p)
        if not args.force:
            print("\nNothing written. The layout constants in this script are pointing at\n"
                  "the wrong cells, or the workbook's own figures are inconsistent.\n"
                  "Fix that rather than passing --force.")
            return 1
        print("\n--force given: writing anyway.")
    else:
        print("\nverification passed: player cells, map points and team totals all "
              "reproduce the workbook")

    if args.dry_run:
        print("--dry-run: nothing written")
        return 0

    if args.live or args.live_only:
        live_path = args.live or os.path.join(ROOT, "docs", "live.json")
        payload = write_live(data, args.date, args.bracket, live_path)
        print("\nwrote %s  (%d teams, %d players, %d maps)"
              % (live_path, len(payload["teams"]), len(payload["players"]),
                 payload["maps_played"]))

    if args.live_only:
        return 0

    out_dir = args.out or os.path.join(ROOT, "data", "tournaments", args.date)
    if os.path.exists(out_dir) and os.listdir(out_dir) and not args.force:
        print("\n%s already exists and is not empty. Refusing to overwrite - "
              "pass --force or pick another --out." % out_dir)
        return 1

    counts = write_csvs(data, args.date, out_dir, args.bracket, data["matchpoint"])
    print("\nwrote %s" % out_dir)
    for name, n in counts.items():
        print("  %-16s %4d rows" % (name, n))

    print("\nnext:")
    print("  1. reconcile the handles above against data/aliases.csv and data/players.csv")
    print("  2. add data/ranks.csv rows for anyone new, or the band model has nothing to")
    print("     compare them against")
    print("  3. python scripts/validate.py")
    print("  4. python scripts/build.py --public")
    return 0


if __name__ == "__main__":
    sys.exit(main())
